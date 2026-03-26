# pyre-ignore-all-errors
from __future__ import annotations

import csv
import io
import logging
import os
import tempfile
import time
from pathlib import Path
from typing import Iterable, List, Optional

from google import genai  # type: ignore
from google.genai import types as genai_types  # type: ignore
from groq import Groq  # type: ignore
from pinecone import Pinecone  # type: ignore
from langchain_text_splitters import RecursiveCharacterTextSplitter  # type: ignore
from pypdf import PdfReader  # type: ignore
import docx  # type: ignore

from django.conf import settings  # type: ignore

from .models import Document, DocumentChunk  # type: ignore


logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".txt", ".xlsx", ".xls", ".csv"}
MAX_RETRIES = 3
RETRY_BACKOFF = 2  # seconds

# Module-level cached clients (lazy-initialized)
_pinecone_index = None
_gemini_client = None
_groq_client = None


def _retry(func, *args, retries=MAX_RETRIES, **kwargs):  # type: ignore
    """Retry a function with exponential backoff."""
    for attempt in range(retries):
        try:
            return func(*args, **kwargs)  # type: ignore
        except Exception as e:
            if attempt == retries - 1:
                logger.error(f"Failed after {retries} attempts: {e}")
                raise
            wait = RETRY_BACKOFF * (2 ** attempt)
            logger.warning(f"Attempt {attempt + 1} failed: {e}. Retrying in {wait}s...")
            time.sleep(wait)


def _ensure_api_keys() -> None:
    if not os.getenv("GEMINI_API_KEY"):
        raise RuntimeError("GEMINI_API_KEY is not set")
    if not os.getenv("GROQ_API_KEY"):
        raise RuntimeError("GROQ_API_KEY is not set")
    if not settings.PINECONE_API_KEY:  # type: ignore
        raise RuntimeError("PINECONE_API_KEY is not set")


def _get_pinecone_index():  # type: ignore
    global _pinecone_index
    if _pinecone_index is None:
        pc = Pinecone(api_key=settings.PINECONE_API_KEY)  # type: ignore
        _pinecone_index = pc.Index(settings.PINECONE_INDEX_NAME)  # type: ignore
    return _pinecone_index


def _get_gemini_client():  # type: ignore
    global _gemini_client
    if _gemini_client is None:
        _gemini_client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))  # type: ignore
    return _gemini_client


def _get_groq_client():  # type: ignore
    global _groq_client
    if _groq_client is None:
        _groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))  # type: ignore
    return _groq_client


def _extract_text(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {suffix}")

    if suffix == ".pdf":
        reader = PdfReader(str(file_path))  # type: ignore
        return "\n".join(page.extract_text() or "" for page in reader.pages)  # type: ignore
    if suffix in (".docx", ".doc"):
        doc = docx.Document(str(file_path))  # type: ignore
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)  # type: ignore
    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(str(file_path), data_only=True)  # type: ignore
            lines = []
            for sheet in wb.sheetnames:  # type: ignore
                ws = wb[sheet]  # type: ignore
                lines.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):  # type: ignore
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)
        except ImportError:
            return file_path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".csv":
        text = file_path.read_text(encoding="utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        return "\n".join("\t".join(row) for row in reader)

    return file_path.read_text(encoding="utf-8", errors="ignore")


def _chunk_text(text: str) -> List[str]:
    splitter = RecursiveCharacterTextSplitter(  # type: ignore
        chunk_size=settings.CHUNK_SIZE,  # type: ignore
        chunk_overlap=settings.CHUNK_OVERLAP,  # type: ignore
        separators=["\n\n", "\n", " ", ""],
    )
    chunks = [chunk.strip() for chunk in splitter.split_text(text) if chunk.strip()]  # type: ignore
    logger.info(f"Split text into {len(chunks)} chunks")
    return chunks


def _embed_texts(texts: Iterable[str]) -> List[List[float]]:
    client = _get_gemini_client()  # type: ignore
    model_name = settings.GEMINI_EMBEDDING_MODEL.removeprefix("models/")  # type: ignore
    dims = getattr(settings, "GEMINI_EMBEDDING_DIMENSIONS", 768)  # type: ignore
    text_list = list(texts)

    if not text_list:
        return []

    # Batch embed — up to 100 texts per API call (Gemini limit)
    all_embeddings: List[List[float]] = []
    batch_size = 100
    for i in range(0, len(text_list), batch_size):
        batch = list(text_list)[i:i + batch_size]  # type: ignore
        logger.info(f"Embedding batch {i // batch_size + 1} ({len(batch)} texts)")
        response = _retry(  # type: ignore
            client.models.embed_content,  # type: ignore
            model=model_name,
            contents=batch,
            config=genai_types.EmbedContentConfig(  # type: ignore
                task_type="RETRIEVAL_DOCUMENT",
                output_dimensionality=dims,
            ),
        )
        for emb in response.embeddings:  # type: ignore
            all_embeddings.append(list(emb.values))  # type: ignore

    logger.info(f"Embedded {len(all_embeddings)} texts total")
    return all_embeddings


def process_document(document: Document) -> None:  # type: ignore
    logger.info(f"Processing document: {document.original_name} (ID: {document.id})")
    _ensure_api_keys()

    # Download file to a temp file — works with both local and S3 storage
    suffix = Path(document.original_name).suffix
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False, mode='wb') as tmp:
            tmp_path = Path(tmp.name)
            # Properly open file from storage (S3 or local)
            try:
                document.file.open('rb')  # type: ignore
                file_content = document.file.read()  # type: ignore
                document.file.close()  # type: ignore
            except Exception as e:
                logger.warning(f"Failed to open file via storage: {e}, trying direct read")
                file_content = document.file.read()  # type: ignore
            
            if not file_content:
                raise ValueError(f"File is empty or could not be read: {document.original_name}")
            
            tmp.write(file_content)
            logger.info(f"Downloaded {len(file_content)} bytes to temp file: {tmp_path}")

        text = _extract_text(tmp_path)
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)  # Clean up temp file

    chunks = _chunk_text(text)

    index = _get_pinecone_index()  # type: ignore
    embeddings = _embed_texts(chunks)

    vectors = []
    for i, chunk in enumerate(chunks):
        vectors.append({
            "id": f"{document.id}-{i}",
            "values": embeddings[i],
            "metadata": {
                "document_id": str(document.id),
                "document_name": document.original_name,
                "chunk_index": i,
                "text": chunk
            }
        })

    # Upsert in batches of 100 to avoid Pinecone limits
    batch_size = 100
    for i in range(0, len(vectors), batch_size):
        _retry(index.upsert, vectors=vectors[i:i + batch_size])  # type: ignore

    DocumentChunk.objects.filter(document=document).delete()  # type: ignore
    DocumentChunk.objects.bulk_create(  # type: ignore
        [
            DocumentChunk(document=document, chunk_index=idx, content=chunk)  # type: ignore
            for idx, chunk in enumerate(chunks)
        ]
    )
    logger.info(f"Document processed successfully: {document.original_name}")


def delete_document_chunks(document_id: int) -> None:
    """Delete all chunks for a document from Pinecone and the database."""
    logger.info(f"Deleting chunks for document ID: {document_id}")
    try:
        index = _get_pinecone_index()  # type: ignore
        chunks = DocumentChunk.objects.filter(document_id=document_id)  # type: ignore
        ids_to_delete = [f"{document_id}-{chunk.chunk_index}" for chunk in chunks]  # type: ignore

        if ids_to_delete:
            batch_size = 100
            for i in range(0, len(ids_to_delete), batch_size):
                _retry(index.delete, ids=ids_to_delete[i:i + batch_size])  # type: ignore
    except Exception as e:
        logger.error(f"Error deleting from Pinecone: {e}")

    DocumentChunk.objects.filter(document_id=document_id).delete()  # type: ignore
    logger.info(f"Chunks deleted for document ID: {document_id}")


def _build_prompt(question: str, context_chunks: List[str], chat_history: Optional[List[dict]] = None) -> List[dict]:
    context_text = "\n\n".join(context_chunks)
    system_prompt = (
        "You are Verba, an expert document Q&A assistant. "
        "Answer ONLY using information from the provided document context. "
        "Do NOT use any outside knowledge.\n\n"
        "Response guidelines:\n"
        "- Use clear structure with headings, bullet points, and numbered lists where helpful\n"
        "- Bold key terms and important information using **bold**\n"
        "- Be concise but thorough\n"
        "- If the context contains tables or data, present them clearly\n"
        "- If the answer spans multiple topics, organize with clear sections\n"
        "- If the document context does not contain sufficient information, say: "
        "'I could not find the answer in your uploaded documents.'\n"
        "- Never fabricate information not present in the context"
    )
    messages = [{"role": "system", "content": system_prompt}]

    if chat_history:
        recent = list(chat_history)[-6:]  # type: ignore
        for msg in recent:
            messages.append({"role": msg["role"], "content": msg["content"]})

    user_prompt = f"Document Context:\n{context_text}\n\nQuestion: {question}"
    messages.append({"role": "user", "content": user_prompt})
    return messages


def answer_question(question: str, document_ids: Optional[List[int]] = None, chat_history: Optional[List[dict]] = None, model: Optional[str] = None) -> dict:  # type: ignore
    logger.info("Answering question: %s...", question[:100])  # type: ignore
    _ensure_api_keys()

    index = _get_pinecone_index()  # type: ignore
    query_embedding = _embed_texts([question])[0]

    filter_dict = None
    if document_ids:
        str_ids = [str(did) for did in document_ids]
        filter_dict = {"document_id": {"$in": str_ids}}

    results = _retry(  # type: ignore
        index.query,  # type: ignore
        vector=query_embedding,
        top_k=5,
        filter=filter_dict,
        include_metadata=True
    )

    documents = []
    metadatas = []

    matches = getattr(results, 'matches', None) or results.get('matches', []) if isinstance(results, dict) else getattr(results, 'matches', [])  # type: ignore
    for match in matches:  # type: ignore
        metadata = getattr(match, 'metadata', None) or (match.get('metadata', {}) if isinstance(match, dict) else {})  # type: ignore
        text = metadata.get('text', '') if isinstance(metadata, dict) else getattr(metadata, 'text', '')  # type: ignore
        documents.append(text)
        metadatas.append(dict(metadata) if not isinstance(metadata, dict) else metadata)  # type: ignore

    # Use the selected model or fall back to default
    selected_model = model if model and model in settings.AVAILABLE_MODELS else settings.GROQ_MODEL  # type: ignore
    logger.info(f"Using model: {selected_model}")

    client = _get_groq_client()  # type: ignore
    response = _retry(  # type: ignore
        client.chat.completions.create,  # type: ignore
        model=selected_model,
        messages=_build_prompt(question, documents, chat_history),
        temperature=0.2,
    )

    answer = response.choices[0].message.content  # type: ignore
    logger.info(f"Answer generated successfully ({len(answer)} chars)")

    sources = []
    for doc_text, metadata in zip(documents, metadatas):
        sources.append(
            {
                "document_id": metadata.get("document_id"),  # type: ignore
                "document_name": metadata.get("document_name"),  # type: ignore
                "chunk_index": metadata.get("chunk_index"),  # type: ignore
                "content": doc_text,
            }
        )

    return {"answer": answer, "sources": sources}


def stream_answer_question(question: str, document_ids: Optional[List[int]] = None, chat_history: Optional[List[dict]] = None, model: Optional[str] = None):  # type: ignore
    """Generator that yields answer tokens as they arrive from Groq streaming API."""
    logger.info("Streaming answer for: %s...", question[:100])  # type: ignore
    _ensure_api_keys()

    index = _get_pinecone_index()  # type: ignore
    query_embedding = _embed_texts([question])[0]

    filter_dict = None
    if document_ids:
        str_ids = [str(did) for did in document_ids]
        filter_dict = {"document_id": {"$in": str_ids}}

    results = _retry(  # type: ignore
        index.query,  # type: ignore
        vector=query_embedding,
        top_k=5,
        filter=filter_dict,
        include_metadata=True
    )

    documents = []
    metadatas = []
    matches = getattr(results, 'matches', None) or results.get('matches', []) if isinstance(results, dict) else getattr(results, 'matches', [])  # type: ignore
    for match in matches:  # type: ignore
        metadata = getattr(match, 'metadata', None) or (match.get('metadata', {}) if isinstance(match, dict) else {})  # type: ignore
        text = metadata.get('text', '') if isinstance(metadata, dict) else getattr(metadata, 'text', '')  # type: ignore
        documents.append(text)
        metadatas.append(dict(metadata) if not isinstance(metadata, dict) else metadata)  # type: ignore

    selected_model = model if model and model in settings.AVAILABLE_MODELS else settings.GROQ_MODEL  # type: ignore
    logger.info(f"Streaming with model: {selected_model}")

    client = _get_groq_client()  # type: ignore
    stream = client.chat.completions.create(  # type: ignore
        model=selected_model,
        messages=_build_prompt(question, documents, chat_history),
        temperature=0.2,
        stream=True,
    )

    full_answer = []
    for chunk in stream:  # type: ignore
        delta = chunk.choices[0].delta  # type: ignore
        if delta and delta.content:  # type: ignore
            full_answer.append(delta.content)  # type: ignore
            yield {"type": "token", "content": delta.content}  # type: ignore

    # Build sources
    sources = []
    for doc_text, metadata in zip(documents, metadatas):
        sources.append({
            "document_id": metadata.get("document_id"),  # type: ignore
            "document_name": metadata.get("document_name"),  # type: ignore
            "chunk_index": metadata.get("chunk_index"),  # type: ignore
            "content": doc_text,
        })

    yield {"type": "done", "answer": "".join(full_answer), "sources": sources}
